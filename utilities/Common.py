import openpyxl
import allure
from allure_commons.types import AttachmentType


class Common:
    url = "https://opensource-demo.orangehrmlive.com/"

    username = "admin"
    password = "admin123"
    title = 'OrangeHRM'
    dash = "Dashboard"
    inv_msg = "Invalid credentials"

    def return_excelvals(self):
        wb = openpyxl.load_workbook("E:\\pytest-allure\\utilities\\testdata.xlsx")
        sh = wb.active
        dict1 = {}
        list_vals = []

        print(sh.max_row)
        print(sh.max_column)

        for ele in range(2, sh.max_row + 1):
            for ele1 in range(1, sh.max_column + 1):
                dict1[sh.cell(row=1, column=ele1).value] = sh.cell(row=ele, column=ele1).value

            list_vals.append(dict1)
            dict1 = {}
        print(list_vals)
        return list_vals

    def return_indv_vals(self):
        lis = self.return_excelvals()

        for i in range(len(lis)):
            return lis[i]

    def takescreenshot(self, tcname, driver):

        allure.attach(driver.get_screenshot_as_png(), name=tcname, attachment_type=AttachmentType.PNG)

